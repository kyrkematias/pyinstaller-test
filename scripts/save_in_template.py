from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def save_in_template(data, excel_path) :

    # Guardar los datos en la primera fila libre del archivo "COPIA PLANTILLA"

    # Cargar el archivo de Excel existente
    excel_path = 'COPIA PLANTILLA.xlsx'
    wb = load_workbook(excel_path)

    # Seleccionar la hoja de cálculo
    sheet = wb.active

    # Obtener la última fila utilizada en la hoja de cálculo
    last_row = sheet.min_row

    # Obtener la próxima fila disponible en la hoja de cálculo
    next_row = last_row + 1

    # Verificar si la próxima fila está vacía
    if all(sheet.cell(row=next_row, column=col).value is None for col in range(1, sheet.max_column + 1)):
        # Agregar los datos en la próxima fila disponible
        for i, value in enumerate(data):
            column_index = i + 1
            column_letter = get_column_letter(column_index)
            cell = f"{column_letter}{next_row}"
            sheet[cell] = value
    else:
        # Buscar la siguiente fila vacía
        while not all(sheet.cell(row=next_row, column=col).value is None for col in range(1, sheet.max_column + 1)):
            next_row += 1

        # Agregar los datos en la siguiente fila vacía
        for i, value in enumerate(data):
            column_index = i + 1
            column_letter = get_column_letter(column_index)
            cell = f"{column_letter}{next_row}"
            sheet[cell] = value

    # Guardar los cambios en el archivo Excel
    wb.save(excel_path)

    print("Datos guardados en COPIA PLANTILLA.xlsx")